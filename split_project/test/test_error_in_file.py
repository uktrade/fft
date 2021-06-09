import os

from openpyxl import Workbook
from split_project.split_figure import PAY_CODE

from split_project.import_project_percentage import (
    WORKSHEET_PROJECT_TITLE,
    COST_CENTRE_CODE,
    NAC_CODE,
    PROGRAMME_CODE,
    ANALYSIS1_CODE,
    ANALYSIS2_CODE,
    PROJECT_CODE,
)

from upload_file.models import FileUpload

from split_project.import_project_percentage import upload_project_percentage_from_file
from split_project.test.test_utils import SplitDataSetup

COST_CENTRE_CODE_INDEX = 1
NAC_CODE_INDEX = 2
PROGRAMME_CODE_INDEX = 3
ANALYSIS1_CODE_INDEX = 4
ANALYSIS2_CODE_INDEX = 5
PROJECT_CODE_INDEX = 6
MONTH1_INDEX = 7
MONTH2_INDEX = 8


class ImportPercentageTest(SplitDataSetup):
    def setUp(self):
        super().setUp()

    def tearDown(self):
        if os.path.exists(self.excel_file_name):
            os.remove(self.excel_file_name)

    def create_workbook(self, data_dictionary):
        wb = Workbook()
        self.data_worksheet = wb.active
        self.data_worksheet.title = WORKSHEET_PROJECT_TITLE
        self.data_worksheet.cell(
            column=COST_CENTRE_CODE_INDEX, row=1, value=COST_CENTRE_CODE
        )
        self.data_worksheet.cell(column=NAC_CODE_INDEX, row=1, value=NAC_CODE)
        self.data_worksheet.cell(
            column=PROGRAMME_CODE_INDEX, row=1, value=PROGRAMME_CODE
        )
        self.data_worksheet.cell(column=PROJECT_CODE_INDEX, row=1, value=PROJECT_CODE)
        self.data_worksheet.cell(
            column=ANALYSIS1_CODE_INDEX, row=1, value=ANALYSIS1_CODE
        )
        self.data_worksheet.cell(
            column=ANALYSIS2_CODE_INDEX, row=1, value=ANALYSIS2_CODE
        )
        self.data_worksheet.cell(column=MONTH1_INDEX, row=1, value="May")
        self.data_worksheet.cell(column=MONTH2_INDEX, row=1, value="Jun")
        row = 2
        for data_row in data_dictionary:
            for data_col, data_value in data_row.items():
                self.data_worksheet.cell(column=data_col, row=row, value=data_value)
            row += 1

        self.excel_file_name = os.path.join(os.path.dirname(__file__), "dummy.xlsx",)
        wb.save(filename=self.excel_file_name)

    def test_wrong_nac_type(self):
        data_dictionary = [
            {
                COST_CENTRE_CODE_INDEX: self.cost_centre_code,
                NAC_CODE_INDEX: self.natural_account_code_non_pay,
                PROGRAMME_CODE_INDEX: self.programme_code,
                PROJECT_CODE_INDEX: self.project_code1,
                MONTH1_INDEX: 0.123,
            },
        ]
        self.create_workbook(data_dictionary)

        file_upload_obj = FileUpload(
            document_file_name=self.excel_file_name,
            document_type=FileUpload.PROJECT_PERCENTAGE,
            file_location=FileUpload.LOCALFILE,
        )
        file_upload_obj.save()
        upload_project_percentage_from_file(self.data_worksheet, file_upload_obj)
        assert (
            f"Row 2 error: The budget category of "
            f"'{self.natural_account_code_non_pay}' is not '{PAY_CODE}'"
            in file_upload_obj.user_error_message
        )

    def test_multiple_directorate(self):
        data_dictionary = [
            {
                COST_CENTRE_CODE_INDEX: self.cost_centre_code,
                NAC_CODE_INDEX: self.natural_account_code_pay,
                PROGRAMME_CODE_INDEX: self.programme_code,
                PROJECT_CODE_INDEX: self.project_code1,
                MONTH1_INDEX: 0.50,
            },
            {
                COST_CENTRE_CODE_INDEX: self.cost_centre_code_different_directorate,
                NAC_CODE_INDEX: self.natural_account_code_pay,
                PROGRAMME_CODE_INDEX: self.programme_code,
                PROJECT_CODE_INDEX: self.project_code1,
                MONTH1_INDEX: 0.30,
            },
        ]
        self.create_workbook(data_dictionary)

        file_upload_obj = FileUpload(
            document_file_name=self.excel_file_name,
            document_type=FileUpload.PROJECT_PERCENTAGE,
            file_location=FileUpload.LOCALFILE,
        )
        file_upload_obj.save()
        upload_project_percentage_from_file(self.data_worksheet, file_upload_obj)

        assert (
            f"Row 3 error: Cost centre "
            f"'{self.cost_centre_code_different_directorate}' is not part"
            in file_upload_obj.user_error_message
        )

    def test_negative_percentage(self):
        data_dictionary = [
            {
                COST_CENTRE_CODE_INDEX: self.cost_centre_code,
                NAC_CODE_INDEX: self.natural_account_code_pay,
                PROGRAMME_CODE_INDEX: self.programme_code,
                PROJECT_CODE_INDEX: self.project_code1,
                MONTH1_INDEX: -0.50,
            },
        ]
        self.create_workbook(data_dictionary)

        file_upload_obj = FileUpload(
            document_file_name=self.excel_file_name,
            document_type=FileUpload.PROJECT_PERCENTAGE,
            file_location=FileUpload.LOCALFILE,
        )
        file_upload_obj.save()
        upload_project_percentage_from_file(self.data_worksheet, file_upload_obj)

        assert (
            "Row 2 error: Negative value in cell" in file_upload_obj.user_error_message
        )

    def test_percentage_too_high(self):
        data_dictionary = [
            {
                COST_CENTRE_CODE_INDEX: self.cost_centre_code,
                NAC_CODE_INDEX: self.natural_account_code_pay,
                PROGRAMME_CODE_INDEX: self.programme_code,
                PROJECT_CODE_INDEX: self.project_code1,
                MONTH1_INDEX: 1.1,
            },
        ]
        self.create_workbook(data_dictionary)

        file_upload_obj = FileUpload(
            document_file_name=self.excel_file_name,
            document_type=FileUpload.PROJECT_PERCENTAGE,
            file_location=FileUpload.LOCALFILE,
        )
        file_upload_obj.save()
        upload_project_percentage_from_file(self.data_worksheet, file_upload_obj)

        assert (
            "Row 2 error: Value higher than 100% in cell"
            in file_upload_obj.user_error_message
        )

    def test_total_percentage_too_high(self):
        data_dictionary = [
            {
                COST_CENTRE_CODE_INDEX: self.cost_centre_code,
                NAC_CODE_INDEX: self.natural_account_code_pay,
                PROGRAMME_CODE_INDEX: self.programme_code,
                PROJECT_CODE_INDEX: self.project_code1,
                MONTH1_INDEX: 0.7,
            },
            {
                COST_CENTRE_CODE_INDEX: self.cost_centre_code,
                NAC_CODE_INDEX: self.natural_account_code_pay,
                PROGRAMME_CODE_INDEX: self.programme_code,
                PROJECT_CODE_INDEX: self.project_code2,
                MONTH1_INDEX: 0.35,
            },
        ]
        self.create_workbook(data_dictionary)

        file_upload_obj = FileUpload(
            document_file_name=self.excel_file_name,
            document_type=FileUpload.PROJECT_PERCENTAGE,
            file_location=FileUpload.LOCALFILE,
        )
        file_upload_obj.save()
        upload_project_percentage_from_file(self.data_worksheet, file_upload_obj)

        assert (
            "The sum of the percentages is higher than 100%"
            in file_upload_obj.user_error_message
        )

    def test_total_percentage_too_low(self):
        data_dictionary = [
            {
                COST_CENTRE_CODE_INDEX: self.cost_centre_code,
                NAC_CODE_INDEX: self.natural_account_code_pay,
                PROGRAMME_CODE_INDEX: self.programme_code,
                PROJECT_CODE_INDEX: self.project_code1,
                MONTH1_INDEX: 0.7,
            },
            {
                COST_CENTRE_CODE_INDEX: self.cost_centre_code,
                NAC_CODE_INDEX: self.natural_account_code_pay,
                PROGRAMME_CODE_INDEX: self.programme_code,
                PROJECT_CODE_INDEX: self.project_code2,
                MONTH1_INDEX: 0.2,
            },
        ]
        self.create_workbook(data_dictionary)

        file_upload_obj = FileUpload(
            document_file_name=self.excel_file_name,
            document_type=FileUpload.PROJECT_PERCENTAGE,
            file_location=FileUpload.LOCALFILE,
        )
        file_upload_obj.save()
        upload_project_percentage_from_file(self.data_worksheet, file_upload_obj)

        assert (
            "The sum of the percentages is lower than 100%"
            in file_upload_obj.user_error_message
        )
