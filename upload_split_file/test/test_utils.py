import os

from openpyxl import Workbook

from chartofaccountDIT.test.factories import (
    ExpenditureCategoryFactory,
    NaturalCodeFactory,
    ProgrammeCodeFactory,
    ProjectCodeFactory,
)
from chartofaccountDIT.models import (
    NaturalCode,
    ProgrammeCode,
    ProjectCode,
)

from core.models import FinancialYear
from core.test.test_base import BaseTestCase, TEST_COST_CENTRE
from core.utils.generic_helpers import make_financial_year_current

from costcentre.test.factories import (
    CostCentreFactory,
    DirectorateFactory,
)
from costcentre.models import CostCentre

from forecast.models import (
    FinancialCode,
    FinancialPeriod,
    ForecastMonthlyFigure,
)
from forecast.utils.import_helpers import VALID_ECONOMIC_CODE_LIST


from upload_split_file.import_project_percentage import (
    WORKSHEET_PROJECT_TITLE,
    COST_CENTRE_CODE,
    NAC_CODE,
    PROGRAMME_CODE,
    ANALYSIS1_CODE,
    ANALYSIS2_CODE,
    PROJECT_CODE,
)

from upload_split_file.models import PaySplitCoefficient

from upload_split_file.split_actuals import INCOME_PAY_CODE, PAY_CODE

COST_CENTRE_CODE_INDEX = 1
NAC_CODE_INDEX = 2
PROGRAMME_CODE_INDEX = 3
ANALYSIS1_CODE_INDEX = 4
ANALYSIS2_CODE_INDEX = 5
PROJECT_CODE_INDEX = 6
MONTH1_INDEX = 7
MONTH2_INDEX = 8


def create_financial_code(
    cost_centre, nac, programme_code, project_code,
):
    programme_obj = ProgrammeCode.objects.get(pk=programme_code)
    costcentre_obj = CostCentre.objects.get(pk=cost_centre)
    nac_obj = NaturalCode.objects.get(pk=nac)
    if project_code:
        project_obj = ProjectCode.objects.get(pk=project_code)
    else:
        project_obj = None
    financial_code_obj, _ = FinancialCode.objects.get_or_create(
        programme=programme_obj,
        cost_centre=costcentre_obj,
        natural_account_code=nac_obj,
        project_code=project_obj,
        analysis1_code=None,
        analysis2_code=None,
    )
    financial_code_obj.save()
    return financial_code_obj


def create_monthly_amount(
    cost_centre, nac, programme_code, project_code, monthly_amount, period_obj,
):
    financial_code_obj = create_financial_code(
        cost_centre, nac, programme_code, project_code,
    )
    forecast_figure_obj = ForecastMonthlyFigure.objects.create(
        financial_year=FinancialYear.objects.get(current=True),
        financial_period=period_obj,
        financial_code=financial_code_obj,
        amount=monthly_amount,
        starting_amount=monthly_amount,
    )
    return forecast_figure_obj


def create_split_data(
    cost_centre, nac, programme_code, project_code, coefficient, period_obj
):
    financial_code_obj = create_financial_code(
        cost_centre, nac, programme_code, project_code,
    )
    directorate_code = financial_code_obj.cost_centre.directorate.directorate_code
    project_split_obj = PaySplitCoefficient.objects.create(
        financial_period=period_obj,
        financial_code_to=financial_code_obj,
        directorate_code=directorate_code,
        split_coefficient=coefficient,
    )
    project_split_obj.save()
    return project_split_obj


def create_workbook(data_dictionary):
    wb = Workbook()
    data_worksheet = wb.active
    data_worksheet.title = WORKSHEET_PROJECT_TITLE
    data_worksheet.cell(column=COST_CENTRE_CODE_INDEX, row=1, value=COST_CENTRE_CODE)
    data_worksheet.cell(column=NAC_CODE_INDEX, row=1, value=NAC_CODE)
    data_worksheet.cell(column=PROGRAMME_CODE_INDEX, row=1, value=PROGRAMME_CODE)
    data_worksheet.cell(column=PROJECT_CODE_INDEX, row=1, value=PROJECT_CODE)
    data_worksheet.cell(column=ANALYSIS1_CODE_INDEX, row=1, value=ANALYSIS1_CODE)
    data_worksheet.cell(column=ANALYSIS2_CODE_INDEX, row=1, value=ANALYSIS2_CODE)
    data_worksheet.cell(column=MONTH1_INDEX, row=1, value="May")
    data_worksheet.cell(column=MONTH2_INDEX, row=1, value="Jun")
    row = 2
    for data_row in data_dictionary:
        for data_col, data_value in data_row.items():
            data_worksheet.cell(column=data_col, row=row, value=data_value)
        row += 1

    excel_file_name = os.path.join(os.path.dirname(__file__), "dummy.xlsx",)
    wb.save(filename=excel_file_name)
    return data_worksheet, excel_file_name


class SplitDataSetup(BaseTestCase):
    def setUp(self):
        self.client.force_login(self.test_user)
        self.test_year = 2019
        make_financial_year_current(self.test_year)
        self.test_period = 9

        self.cost_centre_code = TEST_COST_CENTRE
        self.cost_centre_code1 = TEST_COST_CENTRE + 1
        self.cost_centre_code2 = TEST_COST_CENTRE + 2
        self.cost_centre_code_different_directorate = 234567

        self.natural_account_code_pay = 52191003
        self.natural_account_code_pay1 = 52191004
        self.natural_account_code_pay2 = 52191005
        self.natural_account_code_non_pay = 52191006

        self.programme_code = "310940"
        self.project_code1 = 12341
        self.project_code2 = 12342
        self.project_code3 = 12343
        self.directorate_code = "T123"
        self.directorate_code1 = "T125"
        expenditure_pay_obj = ExpenditureCategoryFactory.create(
            grouping_description=PAY_CODE
        )
        income_pay_obj = ExpenditureCategoryFactory.create(
            grouping_description=INCOME_PAY_CODE
        )

        self.directorate_obj = DirectorateFactory.create(
            directorate_code=self.directorate_code
        )
        directorate_obj1 = DirectorateFactory.create(
            directorate_code=self.directorate_code1
        )
        CostCentreFactory.create(
            cost_centre_code=self.cost_centre_code,
            directorate=self.directorate_obj,
            active=False,
        )
        CostCentreFactory.create(
            cost_centre_code=self.cost_centre_code1,
            directorate=self.directorate_obj,
            active=False,
        )
        CostCentreFactory.create(
            cost_centre_code=self.cost_centre_code2,
            directorate=self.directorate_obj,
            active=False,
        )
        CostCentreFactory.create(
            cost_centre_code=self.cost_centre_code_different_directorate,
            directorate=directorate_obj1,
            active=False,
        )

        NaturalCodeFactory.create(
            natural_account_code=self.natural_account_code_pay,
            economic_budget_code=VALID_ECONOMIC_CODE_LIST[0],
            expenditure_category=income_pay_obj,
            active=False,
        )
        NaturalCodeFactory.create(
            natural_account_code=self.natural_account_code_pay1,
            economic_budget_code=VALID_ECONOMIC_CODE_LIST[0],
            expenditure_category=expenditure_pay_obj,
            active=False,
        )
        NaturalCodeFactory.create(
            natural_account_code=self.natural_account_code_pay2,
            economic_budget_code=VALID_ECONOMIC_CODE_LIST[0],
            expenditure_category=expenditure_pay_obj,
            active=False,
        )
        NaturalCodeFactory.create(
            natural_account_code=self.natural_account_code_non_pay,
            economic_budget_code=VALID_ECONOMIC_CODE_LIST[0],
            active=False,
        )
        ProgrammeCodeFactory.create(
            programme_code=self.programme_code, active=False,
        )
        ProjectCodeFactory.create(project_code=self.project_code1)
        ProjectCodeFactory.create(project_code=self.project_code2)
        ProjectCodeFactory.create(project_code=self.project_code3)

        self.period_obj = FinancialPeriod.objects.get(
            period_calendar_code=self.test_period
        )
