import io

from django.contrib.auth.models import Permission
from django.urls import reverse
from openpyxl import load_workbook

from chartofaccountDIT.test.factories import (
    NaturalCodeFactory,
    ProgrammeCodeFactory,
    ProjectCodeFactory,
)
from core.test.test_base import TEST_COST_CENTRE, BaseTestCase
from core.utils.generic_helpers import (
    get_current_financial_year,
    get_financial_year_obj,
)
from costcentre.test.factories import CostCentreFactory
from forecast.models import BudgetMonthlyFigure, FinancialCode, FinancialPeriod


class DownloadFutureMIBudgetViewTests(BaseTestCase):
    def setUp(self):
        self.client.force_login(self.test_user)
        self.cost_centre_code = TEST_COST_CENTRE
        cost_centre = CostCentreFactory(
            cost_centre_code=self.cost_centre_code,
        )
        self.future_year = get_current_financial_year() + 2
        self.amount_apr = -234567
        self.amount_may = 345216
        self.programme_obj = ProgrammeCodeFactory()
        nac_obj = NaturalCodeFactory()
        project_obj = ProjectCodeFactory()
        self.programme_code = project_obj.project_code
        year_obj = get_financial_year_obj(self.future_year)

        financial_code_obj = FinancialCode.objects.create(
            programme=self.programme_obj,
            cost_centre=cost_centre,
            natural_account_code=nac_obj,
            project_code=project_obj,
        )
        # apr figure
        BudgetMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=1),
            financial_code=financial_code_obj,
            financial_year=year_obj,
            amount=self.amount_apr,
        )
        # may figure
        BudgetMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(
                financial_period_code=2,
            ),
            amount=self.amount_may,
            financial_code=financial_code_obj,
            financial_year=year_obj,
        )

        financial_code_obj1 = FinancialCode.objects.create(
            programme=self.programme_obj,
            cost_centre=cost_centre,
            natural_account_code=nac_obj,
        )
        # another may figure
        BudgetMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(
                financial_period_code=2,
            ),
            amount=self.amount_may,
            financial_code=financial_code_obj1,
            financial_year=year_obj,
        )

        self.year_total = self.amount_apr + self.amount_may

    def test_download_mi_budget(self):
        assert not self.test_user.has_perm("forecast.can_download_mi_reports")
        downloaded_files_url = reverse(
            "download_mi_budget", kwargs={"financial_year": self.future_year}
        )

        response = self.client.get(downloaded_files_url)
        self.assertEqual(response.status_code, 302)
        can_download_files = Permission.objects.get(
            codename="can_download_mi_reports",
        )
        self.test_user.user_permissions.add(can_download_files)
        self.test_user.save()
        response = self.client.get(downloaded_files_url)
        self.assertEqual(response.status_code, 200)
        file = io.BytesIO(response.content)
        wb = load_workbook(
            filename=file,
            read_only=True,
        )
        ws = wb.active
        self.assertEqual(ws["A1"].value, "Entity")
        self.assertEqual(ws["A2"].value, "3000")
        self.assertEqual(ws["I1"].value, "MAY")
        self.assertEqual(ws["H2"].value, self.amount_apr / 100)
        self.assertEqual(ws["I2"].value, self.amount_may / 100)
        self.assertEqual(ws["W2"].value, self.year_total / 100)
        wb.close()
