import io

from bs4 import BeautifulSoup

from django.contrib.auth.models import Permission
from django.urls import reverse

from openpyxl import load_workbook

from chartofaccountDIT.test.factories import (
    NaturalCodeFactory,
    ProgrammeCodeFactory,
    ProjectCodeFactory,
)

from core.models import FinancialYear
from core.test.test_base import BaseTestCase, TEST_COST_CENTRE
from core.utils.generic_helpers import (
    get_current_financial_year,
    get_financial_year_obj,
)

from costcentre.test.factories import CostCentreFactory

from forecast.models import (
    BudgetMonthlyFigure,
    FinancialCode,
    FinancialPeriod,
    ForecastMonthlyFigure,
)


class DownloadViewTests(BaseTestCase):
    def setUp(self):
        self.client.force_login(self.test_user)
        self.cost_centre_code = TEST_COST_CENTRE
        self.cost_centre = CostCentreFactory(cost_centre_code=self.cost_centre_code,)
        current_year = get_current_financial_year()
        self.amount_apr = -234567
        self.amount_may = 345216
        self.programme_obj = ProgrammeCodeFactory()
        nac_obj = NaturalCodeFactory()
        project_obj = ProjectCodeFactory()
        self.programme_code = project_obj.project_code
        year_obj = FinancialYear.objects.get(financial_year=current_year)

        # If you use the MonthlyFigureFactory the test fails.
        # I cannot work out why, it may be due to using a random year....
        financial_code_obj = FinancialCode.objects.create(
            programme=self.programme_obj,
            cost_centre=self.cost_centre,
            natural_account_code=nac_obj,
            project_code=project_obj,
        )
        financial_code_obj.save
        apr_figure = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=1),
            financial_code=financial_code_obj,
            financial_year=year_obj,
            amount=self.amount_apr,
        )
        apr_figure.save
        may_figure = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=2,),
            amount=self.amount_may,
            financial_code=financial_code_obj,
            financial_year=year_obj,
        )
        may_figure.save
        self.year_total = self.amount_apr + self.amount_may

    def test_download_mi_view(self):
        assert not self.test_user.has_perm("forecast.can_download_mi_reports")

        downloaded_files_url = reverse("download_mi_report",)

        # Should have been redirected (no permission)
        resp = self.client.get(downloaded_files_url, follow=False)
        assert resp.status_code == 403

        can_download_files = Permission.objects.get(codename="can_download_mi_reports",)
        self.test_user.user_permissions.add(can_download_files)
        self.test_user.save()

        resp = self.client.get(downloaded_files_url)
        # Should have been permission now
        self.assertEqual(resp.status_code, 200)

    def test_download_mi_report(self):
        assert not self.test_user.has_perm("forecast.can_download_mi_reports")
        downloaded_files_url = reverse(
            "download_mi_report_source",
            kwargs={"financial_year": get_current_financial_year()}
        )

        response = self.client.get(downloaded_files_url)
        self.assertEqual(response.status_code, 302)
        can_download_files = Permission.objects.get(codename="can_download_mi_reports",)
        self.test_user.user_permissions.add(can_download_files)
        self.test_user.save()
        response = self.client.get(downloaded_files_url)
        self.assertEqual(response.status_code, 200)
        file = io.BytesIO(response.content)
        wb = load_workbook(filename=file, read_only=True,)
        ws = wb.active
        self.assertEqual(ws["A1"].value, "Entity")
        self.assertEqual(ws["A2"].value, "3000")
        self.assertEqual(ws["I1"].value, "MAY")
        self.assertEqual(ws["I1"].value, "MAY")
        self.assertEqual(ws["I1"].value, "MAY")
        self.assertEqual(ws["H2"].value, self.amount_apr / 100)
        self.assertEqual(ws["I2"].value, self.amount_may / 100)
        self.assertEqual(ws["W2"].value, self.year_total / 100)
        wb.close()


class DownloadMIBudgetViewTests(BaseTestCase):
    def setUp(self):
        self.client.force_login(self.test_user)
        self.cost_centre_code = TEST_COST_CENTRE
        cost_centre = CostCentreFactory(cost_centre_code=self.cost_centre_code,)
        current_year = get_current_financial_year()
        self.amount_apr = -234567
        self.amount_may = 345216
        self.programme_obj = ProgrammeCodeFactory()
        nac_obj = NaturalCodeFactory()
        project_obj = ProjectCodeFactory()
        self.programme_code = project_obj.project_code
        year_obj = FinancialYear.objects.get(financial_year=current_year)

        # If you use the MonthlyFigureFactory the test fails.
        # I cannot work out why, it may be due to using a random year....
        financial_code_obj = FinancialCode.objects.create(
            programme=self.programme_obj,
            cost_centre=cost_centre,
            natural_account_code=nac_obj,
            project_code=project_obj,
        )
        financial_code_obj.save
        apr_figure = BudgetMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=1),
            financial_code=financial_code_obj,
            financial_year=year_obj,
            amount=self.amount_apr,
        )
        apr_figure.save

        may_figure = BudgetMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=2,),
            amount=self.amount_may,
            financial_code=financial_code_obj,
            financial_year=year_obj,
        )
        may_figure.save

        financial_code_obj1 = FinancialCode.objects.create(
            programme=self.programme_obj,
            cost_centre=cost_centre,
            natural_account_code=nac_obj,
        )

        may_figure1 = BudgetMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=2,),
            amount=self.amount_may,
            financial_code=financial_code_obj1,
            financial_year=year_obj,
        )
        may_figure1.save

        self.year_total = self.amount_apr + self.amount_may

    def test_download_mi_budget(self):
        assert not self.test_user.has_perm("forecast.can_download_mi_reports")
        downloaded_files_url = reverse(
            "download_mi_budget",
            kwargs={"financial_year": get_current_financial_year()}
        )

        response = self.client.get(downloaded_files_url)
        self.assertEqual(response.status_code, 302)
        can_download_files = Permission.objects.get(codename="can_download_mi_reports",)
        self.test_user.user_permissions.add(can_download_files)
        self.test_user.save()
        response = self.client.get(downloaded_files_url)
        self.assertEqual(response.status_code, 200)
        file = io.BytesIO(response.content)
        wb = load_workbook(filename=file, read_only=True,)
        ws = wb.active
        self.assertEqual(ws["A1"].value, "Entity")
        self.assertEqual(ws["A2"].value, "3000")
        self.assertEqual(ws["I1"].value, "MAY")
        self.assertEqual(ws["H2"].value, self.amount_apr / 100)
        self.assertEqual(ws["I2"].value, self.amount_may / 100)
        self.assertEqual(ws["W2"].value, self.year_total / 100)
        wb.close()


class DownloadOscarPermissionTests():
    def setUp(self):
        self.client.force_login(self.test_user)

    def test_download_oscar_view(self):
        assert not self.test_user.has_perm("forecast.can_download_oscar")

        downloaded_files_url = reverse("download_oscar_report", )

        # Should have been redirected (no permission)
        resp = self.client.get(
            downloaded_files_url
        )

        assert resp.status_code == 403

        can_download_files = Permission.objects.get(codename="can_download_oscar", )
        self.test_user.user_permissions.add(can_download_files)
        self.test_user.save()

        resp = self.client.get(downloaded_files_url)

        # Should have been permission now
        self.assertEqual(resp.status_code, 200)


class DownloadMIViewLabelTests(BaseTestCase):
    def setUp(self):
        self.client.force_login(self.test_user)
        can_download_files = Permission.objects.get(codename="can_download_mi_reports",)
        self.test_user.user_permissions.add(can_download_files)
        self.test_user.save()

    def test_previous_year_label(self):
        previous_year_obj = get_financial_year_obj(get_current_financial_year() - 1)
        previous_year_display = previous_year_obj.financial_year_display
        downloaded_files_url = reverse("download_mi_report",)
        response = self.client.get(downloaded_files_url, follow=False)
        soup = BeautifulSoup(response.content, features="html.parser")

        # Check that the text of the button shows the previous year
        previous_year_button = soup.find(id="download_previous_year")
        self.assertIn(previous_year_display, str(previous_year_button))


class DownloadMICurrentYearDropdownTests(BaseTestCase):
    def setUp(self):
        self.client.force_login(self.test_user)
        can_download_files = Permission.objects.get(codename="can_download_mi_reports",)
        self.test_user.user_permissions.add(can_download_files)
        self.test_user.save()

        self.current_year = get_current_financial_year()
        # Create future years
        get_financial_year_obj(self.current_year + 1)
        get_financial_year_obj(self.current_year + 2)
        get_financial_year_obj(self.current_year + 3)

    def test_selectedyear(self):
        downloaded_files_url = reverse("download_mi_report",)
        response = self.client.get(downloaded_files_url, follow=False)

        soup = BeautifulSoup(response.content, features="html.parser")
        download_year_option = soup.find_all('option', selected=True)
        selected_year = download_year_option[0]['value']
        assert (int(selected_year) == self.current_year)
