import io

from django.contrib.auth.models import Permission
from django.urls import reverse

from openpyxl import load_workbook

from previous_years.test.test_utils import (
    PastYearForecastSetup,
)


class DownloadPreviousYearMIReportTest(PastYearForecastSetup):
    def setUp(self):
        self.client.force_login(self.test_user)
        can_view_forecasts = Permission.objects.get(codename="can_download_mi_reports")
        self.test_user.user_permissions.add(can_view_forecasts)
        self.test_user.save()
        super().setUp()

    def test_download(self):
        response = self.client.get(
            reverse("download_mi_previous_year_report_source"),
        )
        self.assertEqual(response.status_code, 200)

        file = io.BytesIO(response.content)
        wb = load_workbook(filename=file)
        ws = wb.active
        assert ws["A1"].value == "Entity"
        assert ws["B1"].value == "Cost Centre"
        assert ws["B2"].value == f"{self.cost_centre_code}"
        assert ws["C1"].value == "Natural Account"
        assert ws["C2"].value == self.natural_account_code
        assert ws["D1"].value == "Programme"
        assert ws["D2"].value == self.programme_code
        assert ws["E1"].value == "Analysis"
        assert ws["E2"].value == self.analisys1
        assert ws["F1"].value == "Analysis2"
        assert ws["F2"].value == self.analisys2
        assert ws["G1"].value == "Project"
        assert ws["G2"].value == self.project_code
        assert ws["W1"].value == "Total"
        assert ws["W2"].value == self.year_total
        assert ws["H1"].value == "APR"
        assert ws["H2"].value == self.outturn["apr"]
        assert ws["I1"].value == "MAY"
        assert ws["I2"].value == self.outturn["may"]
        assert ws["J1"].value == "JUN"
        assert ws["J2"].value == self.outturn["jun"]
        assert ws["K1"].value == "JUL"
        assert ws["K2"].value == self.outturn["jul"]
        assert ws["L1"].value == "AUG"
        assert ws["L2"].value == self.outturn["aug"]
        assert ws["M1"].value == "SEP"
        assert ws["M2"].value == self.outturn["sep"]
        assert ws["N1"].value == "OCT"
        assert ws["N2"].value == self.outturn["oct"]
        assert ws["O1"].value == "NOV"
        assert ws["O2"].value == self.outturn["nov"]
        assert ws["P1"].value == "DEC"
        assert ws["P2"].value == self.outturn["dec"]
        assert ws["Q1"].value == "JAN"
        assert ws["Q2"].value == self.outturn["jan"]
        assert ws["R1"].value == "FEB"
        assert ws["R2"].value == self.outturn["feb"]
        assert ws["S1"].value == "MAR"
        assert ws["S2"].value == self.outturn["mar"]
        assert ws["T1"].value == "ADJ01"
        assert ws["T2"].value == self.outturn["adj01"]
        assert ws["U1"].value == "ADJ02"
        assert ws["U2"].value == self.outturn["adj02"]
        assert ws["V1"].value == "ADJ03"
        assert ws["V2"].value == self.outturn["adj03"]
