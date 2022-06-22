import io

from django.contrib.auth.models import Permission
from django.urls import reverse

from openpyxl import load_workbook

from chartofaccountDIT.test.factories import (
    NaturalCodeFactory,
    ProgrammeCodeFactory,
    ProjectCodeFactory,
)

from core.test.test_base import TEST_COST_CENTRE, BaseTestCase
from core.utils.generic_helpers import (
    get_current_financial_year,
    get_financial_year_obj,
)

from costcentre.test.factories import (
    CostCentreFactory,
    DepartmentalGroupFactory,
    DirectorateFactory,
)

from forecast.models import (
    FinancialCode,
    FinancialPeriod,
    ForecastMonthlyFigure,
)

from forecast.utils.view_header_definition import (
    forecast_total_header,
    year_to_date_header,
)


class DownloadForecastHierarchyTest(BaseTestCase):
    def setUp(self):
        self.client.force_login(self.test_user)
        self.group_name = "Test Group"
        self.group_code = "TestGG"
        directorate_name = "Test Directorate"
        directorate_code = "TestDD"
        cost_centre_code = TEST_COST_CENTRE


        group = DepartmentalGroupFactory(
            group_code=self.group_code,
            group_name=self.group_name,
        )
        directorate = DirectorateFactory(
            directorate_code=directorate_code,
            directorate_name=directorate_name,
            group=group,
        )
        cost_centre = CostCentreFactory(
            directorate=directorate,
            cost_centre_code=cost_centre_code,
        )
        current_year = get_current_financial_year()
        self.next_year = current_year + 1
        programme_obj = ProgrammeCodeFactory()
        nac_obj = NaturalCodeFactory()
        project_obj = ProjectCodeFactory()
        year_obj = get_financial_year_obj(current_year)
        next_year_obj = get_financial_year_obj(self.next_year)
        # Create actuals
        for period in range(1, 4):
            period_obj = FinancialPeriod.objects.get(financial_period_code=period)
            period_obj.actual_loaded = True
            period_obj.save()

        financial_code_obj = FinancialCode.objects.create(
            programme=programme_obj,
            cost_centre=cost_centre,
            natural_account_code=nac_obj,
            project_code=project_obj,
        )
        financial_code_obj.save
        self.amount_apr_current_year = 987654300
        apr_figure = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=1),
            financial_code=financial_code_obj,
            financial_year=year_obj,
            amount=self.amount_apr_current_year,
        )
        apr_figure.save
        self.amount_apr_next_year = 9898989800
        next_year_april_figures = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(
                financial_period_code=1,
            ),
            amount=self.amount_apr_next_year,
            financial_code=financial_code_obj,
            financial_year=next_year_obj,
        )
        next_year_april_figures.save
        # Assign forecast view permission
        can_view_forecasts = Permission.objects.get(codename="can_view_forecasts")
        self.test_user.user_permissions.add(can_view_forecasts)
        self.test_user.save()

    def test_current_year_download(self):
        dit_url = self.client.get(
            reverse("export_forecast_data_dit", kwargs={"period": 0})
        )

        file = io.BytesIO(dit_url.content)
        wb = load_workbook(filename=file)
        ws = wb.active
        # Check group
        assert ws["A1"].value == "Group name"
        assert ws["B2"].value == self.group_code

        assert ws["Z1"].value == "Apr"
        assert ws["Z2"].value == self.amount_apr_current_year / 100
        assert ws["AO1"].value == forecast_total_header
        assert ws["AO1"].value == forecast_total_header
        # Check year to date exists
        assert ws["AQ1"].value == year_to_date_header
        # Formula to calculate year to date
        assert ws["AQ2"].value == "=SUM(Z2:AB2)"

    def test_next_year_download(self):
        dit_url = self.client.get(
            reverse("export_forecast_data_dit", kwargs={"period": self.next_year})
        )

        file = io.BytesIO(dit_url.content)
        wb = load_workbook(filename=file)
        ws = wb.active
        # Check group
        assert ws["A1"].value == "Group name"
        assert ws["B2"].value == self.group_code
        assert ws["Z1"].value == "Apr"
        assert ws["Z2"].value == self.amount_apr_next_year / 100
        assert ws["AO1"].value == forecast_total_header
        assert ws["AO1"].value == forecast_total_header
        # Check year to date does exists
        assert ws["AQ1"].value is None
        assert ws["AQ2"].value is None
