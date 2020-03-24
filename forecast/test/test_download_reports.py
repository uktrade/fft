import io

from django.contrib.auth.models import Permission
from django.test import TestCase
from django.urls import reverse

from openpyxl import load_workbook

from chartofaccountDIT.test.factories import (
    NaturalCodeFactory,
    ProgrammeCodeFactory,
    ProjectCodeFactory,
)

from core.models import FinancialYear
from core.myutils import get_current_financial_year
from core.test.test_base import RequestFactoryBase

from costcentre.test.factories import (
    CostCentreFactory,
)

from forecast.models import (
    FinancialCode,
    FinancialPeriod,
    ForecastMonthlyFigure,
)
from forecast.views.export.mi_report_source import export_mi_report
from forecast.views.export.oscar_return import export_oscar_report

from treasuryCOA.test.factories import L5AccountFactory

class DownloadMIReportTest(TestCase, RequestFactoryBase):
    def setUp(self):
        RequestFactoryBase.__init__(self)
        self.cost_centre_code = 109076
        cost_centre = CostCentreFactory(cost_centre_code=self.cost_centre_code,)
        current_year = get_current_financial_year()
        self.amount_apr = -9876543
        programme_obj = ProgrammeCodeFactory()
        self.programme_code = programme_obj.programme_code
        treasury_l5 =  L5AccountFactory()
        nac_obj = NaturalCodeFactory(account_L5_code=treasury_l5)
        self.nac = nac_obj.natural_account_code
        project_obj = ProjectCodeFactory()
        self.project_code = project_obj.project_code
        year_obj = FinancialYear.objects.get(financial_year=current_year)

        apr_period = FinancialPeriod.objects.get(financial_period_code=1)
        apr_period.actual_loaded = True
        apr_period.save()

        # If you use the MonthlyFigureFactory the test fails.
        # I cannot work out why, it may be due to using a random year....
        financial_code_obj = FinancialCode.objects.create(
            programme=programme_obj,
            cost_centre=cost_centre,
            natural_account_code=nac_obj,
            project_code=project_obj,
        )
        financial_code_obj.save
        apr_figure = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=1),
            financial_code=financial_code_obj,
            financial_year=year_obj,
            amount=self.amount_apr,
        )
        apr_figure.save
        self.amount_may = 1234567
        may_figure = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=2,),
            amount=self.amount_may,
            financial_code=financial_code_obj,
            financial_year=year_obj,
        )
        may_figure.save
        # Assign forecast view permission
        can_view_forecasts = Permission.objects.get(codename="can_view_forecasts")
        self.test_user.user_permissions.add(can_view_forecasts)
        self.test_user.save()

        self.year_total = self.amount_apr + self.amount_may

    def test_download(self):
        response = self.factory_get(
            reverse("download_mi_report_source"), export_mi_report,
        )

        self.assertEqual(response.status_code, 200)

        file = io.BytesIO(response.content)
        wb = load_workbook(filename=file)
        ws = wb.active
        assert ws["A1"].value == "Entity"
        assert ws["B1"].value == "Cost Centre"
        assert ws["B2"].value == f"{self.cost_centre_code}"
        assert ws["C1"].value == "Natural Account"
        assert ws["C2"].value == self.nac
        assert ws["G1"].value == 'Project'
        assert ws["G2"].value == self.project_code
        assert ws["W1"].value == "Total"
        assert ws["W2"].value == self.year_total / 100
        assert ws["H1"].value == "APR"
        assert ws["H2"].value == self.amount_apr / 100
        assert ws["I1"].value == "MAY"
        assert ws["I2"].value == self.amount_may / 100


class DownloadOscarReportTest(TestCase, RequestFactoryBase):
    def setUp(self):
        RequestFactoryBase.__init__(self)
        self.cost_centre_code = 109076
        cost_centre = CostCentreFactory(cost_centre_code=self.cost_centre_code,)
        current_year = get_current_financial_year()
        self.amount_apr = -9876543
        programme_obj = ProgrammeCodeFactory()
        self.programme_code = programme_obj.programme_code
        nac_obj = NaturalCodeFactory()
        self.nac = nac_obj.natural_account_code
        project_obj = ProjectCodeFactory()
        self.project_code = project_obj.project_code
        year_obj = FinancialYear.objects.get(financial_year=current_year)

        apr_period = FinancialPeriod.objects.get(financial_period_code=1)
        apr_period.actual_loaded = True
        apr_period.save()

        # If you use the MonthlyFigureFactory the test fails.
        # I cannot work out why, it may be due to using a random year....
        financial_code_obj = FinancialCode.objects.create(
            programme=programme_obj,
            cost_centre=cost_centre,
            natural_account_code=nac_obj,
            project_code=project_obj,
        )
        financial_code_obj.save
        apr_figure = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=1),
            financial_code=financial_code_obj,
            financial_year=year_obj,
            amount=self.amount_apr,
        )
        apr_figure.save
        self.amount_may = 1234567
        may_figure = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(financial_period_code=2,),
            amount=self.amount_may,
            financial_code=financial_code_obj,
            financial_year=year_obj,
        )
        may_figure.save
        # Assign forecast view permission
        can_view_forecasts = Permission.objects.get(codename="can_view_forecasts")
        self.test_user.user_permissions.add(can_view_forecasts)
        self.test_user.save()

        self.year_total = self.amount_apr + self.amount_may

    def test_download(self):
        response = self.factory_get(
            reverse("download_oscar"),
            export_oscar_report,
        )

        self.assertEqual(response.status_code, 200)

        file = io.BytesIO(response.content)
        wb = load_workbook(filename=file)
        ws = wb.active
        assert ws["A1"].value == 'Row'
        assert ws["B1"].value == 'Organisation'
        assert ws["C1"].value == 'Organisation Alias'
        assert ws["D1"].value == 'COA'
        assert ws["E1"].value == 'COA Alias'
