from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Protection
from openpyxl.utils import column_index_from_string, get_column_letter

from core.utils.export_helpers import EXC_TAB_NAME_LEN, EXCEL_TYPE
from core.utils.generic_helpers import get_current_financial_year, today_string
from end_of_month.utils import monthly_variance_exists
from forecast.models import FinancialPeriod
from forecast.utils.view_field_definition import outturn_variance_field
from forecast.utils.view_header_definition import (
    budget_header,
    forecast_total_header,
    variance_header,
    variance_outturn_header,
    year_to_date_header,
)


def format_numbers(ws, row, start):
    for c in range(start, start + 16):
        ws.cell(column=c, row=row).number_format = "#,##0.00"


def unlock_forecast_cells(ws, row, start, end):
    for c in range(start, end):
        ws.cell(column=c, row=row).protection = Protection(locked=False)


def get_obj_value(obj, name):
    value = 0
    if name in obj:
        value = obj[name]
        if value is None:
            value = 0
    return value


def forecast_query_iterator(
    queryset, keys_dict, columns_dict, period_list, monthly_variance_dict
):
    for obj in queryset:
        row = []
        for field in keys_dict.keys():
            val = obj[field]
            if val is None:
                val = ""
            row.append(val)
        row.append(obj["Budget"] / 100)
        for period in period_list:
            row.append(obj[period] / 100)
        # space holder for outturn
        row.append("")
        for field in monthly_variance_dict.keys():
            row.append(obj[field] / 100)

        # space holder for variance
        row.append("")
        # space holder for variance
        row.append("")
        for field in columns_dict.keys():
            val = obj[field]
            if val is None:
                val = ""
            row.append(val)
        yield row


def create_headers(
    keys_dict, columns_dict, period_list, show_monthly_variance, show_spent_to_date
):
    k = list(keys_dict.values())
    k.append(budget_header)
    k.extend(period_list)
    k.append(forecast_total_header)
    k.extend(list(show_monthly_variance.values()))
    k.append(variance_header)
    if show_spent_to_date:
        k.append(year_to_date_header)
    k.extend(list(columns_dict.values()))
    return k


def export_forecast_to_excel(  # noqa C901
    queryset,
    columns_dict,
    extra_columns_dict,
    protect,
    title,
    include_month_total,
    period_to_show,
    display_monthly_variance,
):
    resp = HttpResponse(content_type=EXCEL_TYPE)
    filename = f"{title}  {today_string()} .xlsx"
    resp["Content-Disposition"] = "attachment; filename=" + filename
    wb = Workbook()
    ws = wb.active
    # Truncate the tab name to the maximum length permitted by Excel
    ws.title = f"{title} {today_string()}"[:EXC_TAB_NAME_LEN]
    if protect:
        #  Set the required level of protection
        ws.protection.sheet = True
        ws.protection.autoFilter = False
        ws.protection.sort = False
        ws.protection.pivotTables = False
        ws.protection.formatCells = False
        ws.protection.formatRows = False
        ws.protection.formatColumns = False
    row_count = 1

    current_year = get_current_financial_year()
    if period_to_show > 2000:
        display_previous_years = period_to_show < current_year
        display_future_years = period_to_show > current_year
        # As the argument had the year in it, the period to show is not defined.
        period_to_show = 0
    else:
        display_previous_years = False
        display_future_years = False

    if display_previous_years:
        # Display adjustments periods when showing datas from previous years
        period_list = FinancialPeriod.financial_period_info.period_display_all_list()
    else:
        period_list = FinancialPeriod.financial_period_info.period_display_list()

    howmany_periods = len(period_list)
    monthly_variance_dict = {}
    if display_monthly_variance:
        monthly_variance_dict = {outturn_variance_field: variance_outturn_header}
    header = create_headers(
        columns_dict,
        extra_columns_dict,
        period_list,
        monthly_variance_dict,
        not display_future_years,
    )
    budget_index = header.index(budget_header) + 1
    budget_col = get_column_letter(budget_index)
    first_figure_index = budget_index + 1
    first_figure_col = get_column_letter(first_figure_index)
    if display_previous_years:
        # For previous years, all the periods are actuals
        howmany_actuals = howmany_periods
    elif display_future_years:
        # No actuals in future forecasts
        howmany_actuals = 0
    else:
        # Actual month starts at 1 for April,
        # so it can be used as counter of the actual periods
        if period_to_show:
            howmany_actuals = period_to_show
        else:
            # download the current period
            howmany_actuals = FinancialPeriod.financial_period_info.actual_month()
    first_forecast_index = first_figure_index + howmany_actuals
    if howmany_actuals:
        last_actual_col = get_column_letter(budget_index + howmany_actuals)
    last_month_index = budget_index + howmany_periods
    last_month_col = get_column_letter(last_month_index)
    next_index = last_month_index + 1
    year_total_col = get_column_letter(next_index)
    if display_monthly_variance:
        next_index += 1
    next_index += 1
    over_under_spend_col = get_column_letter(next_index)
    next_index += 1
    year_to_date_col = get_column_letter(next_index)
    ws.append(header)

    for data_row in forecast_query_iterator(
        queryset, columns_dict, extra_columns_dict, period_list, monthly_variance_dict
    ):
        ws.append(data_row)
        row_count += 1
        # Formula for Year To Date. Don't use it if there are no actuals
        if howmany_actuals:
            ws[f"{year_to_date_col}{row_count}"].value = (
                f"=SUM({first_figure_col}{row_count}:{last_actual_col}{row_count})"
            )
        else:
            if not display_future_years:
                ws[f"{year_to_date_col}{row_count}"].value = 0

        # Formula for calculating the full year
        ws[f"{year_total_col}{row_count}"].value = (
            f"=SUM({first_figure_col}{row_count}:{last_month_col}{row_count})"
        )
        ws[f"{over_under_spend_col}{row_count}"].value = (
            f"=({budget_col}{row_count}-{year_total_col}{row_count})"
        )

        format_numbers(ws, row_count, budget_index)
        unlock_forecast_cells(ws, row_count, first_forecast_index, last_month_index + 1)

    if include_month_total:
        row_count += 1
        grand_total_column = get_column_letter(column_index_from_string(budget_col) - 1)
        ws[f"{grand_total_column}{row_count}"].value = "Grand Total:"

        for col_idx in range(
            column_index_from_string(budget_col),
            column_index_from_string(year_to_date_col) + 1,
        ):
            col = get_column_letter(col_idx)
            ws[f"{col}{row_count}"].value = f"=SUM({col}2:{col}{row_count-1})"

    wb.save(resp)
    return resp


def export_query_to_excel(queryset, columns_dict, title, period):
    show_monthly_variance = monthly_variance_exists(period)
    return export_forecast_to_excel(
        queryset, columns_dict, {}, False, title, False, period, show_monthly_variance
    )


def export_edit_to_excel(queryset, key_dict, columns_dict, title, financial_year):
    return export_forecast_to_excel(
        queryset, key_dict, columns_dict, True, title, True, int(financial_year), False
    )
